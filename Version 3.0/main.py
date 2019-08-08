import glob
import PyPDF2
import os
import pdf
import swg_general as swg
from openpyxl import load_workbook
import openpyxl
from time import gmtime, strftime

curDir = os.getcwd()
print("SWG PDF Extraction Tool\n1. Convert PDF to hOCR\n2. Extract Data from hOCR output to Excel\n3. Exit")
user_input = input("Input: ")
while user_input != "3":
    if user_input == "1":
        valdiate = input("Are you sure this will overwrite previous hOCR: [yes] / [no] ")
        if valdiate.lower() == "yes":
            pdf_test = pdf.PDF_2_HOCR("THIS POLICY CONTAINS A CLAUSE WHICH MAY LIMIT THE AMOUNT PAYABLE",curDir,10,1)
    if user_input == "2":
        os.chdir(curDir + '\\Excel_Output')
        template = load_workbook('Template.xlsx')
        i = 2
        sheet = template.active
        swg_extract = swg.SWG_Extract(curDir + '\\hOCR_OUT')
        excel_rows = []
        for row in swg_extract.borderauxCollection:
            try:
                row.searchTableData()
                row.calculatedFields()
                row.splitAddress()
                excel_rows.append(row.exportToExcel())
            except Exception:
                print("Error in extraction")
                row.show()
        for row in excel_rows:
            for cell in sheet["A"]:
                if cell.value is None:
                    lrow = cell.row
                    break
                else:
                    lrow = cell.row + 1
        
            for loc in row:
                
                for cell in loc:
                    sheet.cell(row = i, column = cell[0]).value = cell[1]
                i+= 1
            '''
            for index, entry in enumerate(row):
                sheet.cell(row = i, column = entry[0]).value = entry[1]
                if index % 2 == 0:
                    i+= 1
            '''
        date = strftime("%Y-%m-%d %H:%M:%S", gmtime()).split()

        os.chdir(curDir + '\\Excel_Output\\Outputs')
        outputs = glob.glob('*.xlsx')
        template.save(filename = curDir + '\\Excel_Output\\Outputs\\' + str(date[0]) + '_' + str(len(outputs)+1) + '.xlsx')
    
    print("SWG PDF Extraction Tool\n1. Convert PDF to hOCR\n2. Extract Data from hOCR output to Excel\n3. Exit")
    user_input = input("Input: ")   
